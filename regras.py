import openpyxl
from openpyxl import Workbook
from pathlib import Path

def extrair_e_salvar_regras_como_texto(nome_aba_alvo="MAR 2026"):
    diretorio_atual = Path(__file__).parent
    # Lista arquivos ignorando temporários e o próprio arquivo de saída
    arquivos_excel = [f for f in diretorio_atual.glob("*.xlsx") 
                      if not f.name.startswith("~$") and "REGRAS_RATEIO_CONSOLIDADO" not in f.name]
    
    if not arquivos_excel:
        print("Nenhum arquivo .xlsx de origem encontrado.")
        return

    dados_finais = []
    vistos = set()

    for caminho_arquivo in arquivos_excel:
        print(f"--- Processando: {caminho_arquivo.name} ---")
        try:
            wb = openpyxl.load_workbook(caminho_arquivo, data_only=False)
            if nome_aba_alvo not in wb.sheetnames:
                print(f"Aviso: Aba '{nome_aba_alvo}' não encontrada. Pulando.")
                wb.close()
                continue
            
            ws = wb[nome_aba_alvo]
            
            # Mapeamento dinâmico de cabeçalhos
            headers = {str(cell.value).strip().upper(): cell.column - 1 for cell in ws[1] if cell.value}
            
            # Definição de índices (com fallback para colunas padrão se não encontrar pelo nome)
            idx_cartao = headers.get("Nº CARTÃO") or headers.get("CARTAO") or 5
            idx_unidade = headers.get("UNIDADES") or 0
            idx_filial = headers.get("FILIAL") or 1
            idx_cr = headers.get("CR") or 2
            idx_projeto = headers.get("PROJETO") or 3
            idx_valor_h = 7 # Coluna H
            idx_valor_i = 8 # Coluna I

            for row in ws.iter_rows(min_row=2, max_col=max(headers.values()) + 1 if headers else 15):
                num_cartao = row[idx_cartao].value
                formula_h = row[idx_valor_h].value
                formula_i = row[idx_valor_i].value

                # Identifica fórmulas
                f_h_str = str(formula_h) if formula_h and str(formula_h).startswith('=') else None
                f_i_str = str(formula_i) if formula_i and str(formula_i).startswith('=') else None

                if num_cartao and (f_h_str or f_i_str):
                    unidade = row[idx_unidade].value
                    filial = row[idx_filial].value
                    cr = row[idx_cr].value
                    projeto = row[idx_projeto].value

                    # Assinatura para remover duplicatas
                    assinatura = (unidade, filial, cr, projeto, num_cartao, f_h_str, f_i_str)

                    if assinatura not in vistos:
                        vistos.add(assinatura)
                        
                        # Adicionamos o apóstrofo (') no início para o Excel não executar a fórmula
                        # O apóstrofo fica invisível na célula, mas força o formato texto
                        f_h_formatada = f"'{f_h_str}" if f_h_str else "---"
                        f_i_formatada = f"'{f_i_str}" if f_i_str else "---"

                        dados_finais.append({
                            "UNIDADES": unidade,
                            "FILIAL": filial,
                            "CR": cr,
                            "PROJETO": projeto,
                            "Nº CARTÃO": num_cartao,
                            "FÓRMULA VALOR (H)": f_h_formatada,
                            "FÓRMULA VSC (I)": f_i_formatada,
                            "ARQUIVO ORIGEM": caminho_arquivo.name
                        })

            wb.close()
        except Exception as e:
            print(f"Erro ao processar {caminho_arquivo.name}: {e}")

    # Geração do arquivo de saída
    if dados_finais:
        novo_wb = Workbook()
        nova_ws = novo_wb.active
        nova_ws.title = "Regras Extraídas"

        colunas = list(dados_finais[0].keys())
        nova_ws.append(colunas)

        for item in dados_finais:
            # Escreve os dados linha por linha
            nova_ws.append([item[col] for col in colunas])

        # Formatação básica de largura de coluna
        for col in nova_ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value: max_length = max(max_length, len(str(cell.value)))
                except: pass
            nova_ws.column_dimensions[column_letter].width = max_length + 3

        nome_saida = diretorio_atual / "REGRAS_RATEIO_CONSOLIDADO.xlsx"
        try:
            novo_wb.save(nome_saida)
            print(f"\n✅ Arquivo gerado com sucesso: {nome_saida.name}")
            print(f"📌 Total de regras únicas encontradas: {len(dados_finais)}")
        except PermissionError:
            print(f"\n❌ Erro: O arquivo '{nome_saida.name}' está aberto. Feche-o e tente novamente.")
    else:
        print("\n❌ Nenhuma regra de rateio encontrada.")

if __name__ == "__main__":
    extrair_e_salvar_regras_como_texto("MAR 2026")