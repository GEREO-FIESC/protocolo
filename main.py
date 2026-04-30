import FreeSimpleGUI as sg
import pandas as pd
import threading
import logging
import re
import os
from datetime import datetime
from pathlib import Path
from openpyxl.styles import PatternFill

# =============================================================================
# CONFIGURAÇÕES TÉCNICAS E LOGGING
# =============================================================================

def configurar_logger():
    """Configura o log em arquivo com timestamp."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_arquivo = f"log_rateio_{timestamp}.txt"
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(nome_arquivo, encoding='utf-8'),
        ]
    )
    return nome_arquivo

# =============================================================================
# LÓGICA DE PROCESSAMENTO (CORE)
# =============================================================================

class ProcessadorRateio:
    def __init__(self, window, caminhos):
        self.window = window
        self.path_extrato = caminhos['-EXTRATO-']
        self.path_inicial = caminhos['-INICIAL-']
        self.path_regras = caminhos['-REGRAS-']
        self.path_saida = Path(self.path_inicial).parent / "CONSOLIDADO_RATEIO_FINAL.xlsx"

    def log_gui(self, mensagem, nivel="INFO"):
        if nivel == "INFO":
            logging.info(mensagem)
        elif nivel == "ERROR":
            logging.error(mensagem)
        
        # Envia evento para a thread principal atualizar a GUI
        self.window.write_event_value("-THREAD_LOG-", f"[{nivel}] {mensagem}")

    def _converter_numerico(self, valor):
        if pd.isna(valor) or str(valor).strip() == "": return 0.0
        if isinstance(valor, (int, float)): return float(valor)
        s = str(valor).replace("R$", "").replace(".", "").replace(",", ".").strip()
        try: return float(s)
        except: return 0.0

    def _limpar_id(self, valor):
        if pd.isna(valor): return ""
        return str(valor).split('.')[0].strip()

    def executar(self):
        try:
            self.log_gui("Iniciando processamento de dados...")
            
            # 1. Carregar Extrato
            self.log_gui("Lendo planilha de extrato...")
            if self.path_extrato.lower().endswith('.csv'):
                df_extrato_raw = pd.read_csv(self.path_extrato, sep=None, engine='python', encoding='latin1', header=None)
            else:
                df_extrato_raw = pd.read_excel(self.path_extrato, header=None)

            dict_extrato = {}
            for _, row in df_extrato_raw.iterrows():
                if "Total Cartão" in str(row[0]):
                    match = re.search(r"Total Cartão (\d+)", str(row[0]))
                    if match:
                        id_cartao = self._limpar_id(match.group(1))
                        dict_extrato[id_cartao] = round(self._converter_numerico(row[7]), 2)

            # 2. Carregar Inicial
            self.log_gui("Lendo planilha base inicial...")
            df = pd.read_excel(self.path_inicial)
            df.columns = [str(c).strip() for c in df.columns]
            
            for col in ['VALOR', 'VALOR VSC PERCURSO', 'TOTAL CONTRATO', 'VALOR TOTAL']:
                if col not in df.columns: df[col] = 0.0
            
            df['Nº CARTÃO'] = df['Nº CARTÃO'].apply(self._limpar_id)
            df['TOTAL CONTRATO'] = df['TOTAL CONTRATO'].apply(lambda x: round(self._converter_numerico(x), 2))
            df['VALOR'] = 0.0
            df['VALOR TOTAL'] = 0.0

            # 3. Rateio e VSC
            self.log_gui("Aplicando regras financeiras e arredondamentos...")
            for cartao, v_total in dict_extrato.items():
                indices = df.index[df['Nº CARTÃO'] == cartao].tolist()
                if not indices: continue

                idx_pai = indices[0]
                idx_filhas = indices[1:]

                df.at[idx_pai, 'VALOR TOTAL'] = v_total
                df.at[idx_pai, 'VALOR'] = v_total

                if idx_filhas:
                    v_rateado = round(v_total / len(idx_filhas), 2)
                    for f_idx in idx_filhas:
                        df.at[f_idx, 'VALOR'] = v_rateado
                    
                    dif = round(v_total - round(v_rateado * len(idx_filhas), 2), 2)
                    if dif != 0: df.at[idx_filhas[-1], 'VALOR'] = round(df.at[idx_filhas[-1], 'VALOR'] + dif, 2)

                for idx in indices:
                    df.at[idx, 'VALOR VSC PERCURSO'] = round(df.at[idx, 'TOTAL CONTRATO'] - df.at[idx, 'VALOR'], 2)

            # 4. Salvar e Colorir
            self.log_gui("Gerando arquivo Excel consolidado...")
            df = df.fillna(0)
            contagem = df['Nº CARTÃO'].value_counts().to_dict()

            with pd.ExcelWriter(self.path_saida, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='CONSOLIDADO')
                ws = writer.sheets['CONSOLIDADO']
                cor = PatternFill(start_color='C4D79B', end_color='C4D79B', fill_type='solid')

                col_nomes = list(df.columns)
                idx_cartao_col = col_nomes.index('Nº CARTÃO') + 1
                cols_filha = [i+1 for i, c in enumerate(col_nomes) if c in ['Nº CARTÃO', 'MALOTE/POSTAGEM', 'VALOR', 'VALOR VSC PERCURSO', 'VALOR TOTAL']]

                ant = None
                for r in range(2, ws.max_row + 1):
                    atual = self._limpar_id(ws.cell(row=r, column=idx_cartao_col).value)
                    if not atual or contagem.get(atual, 0) <= 1: 
                        ant = atual
                        continue
                    
                    if atual != ant:
                        for c in range(1, ws.max_column + 1): ws.cell(row=r, column=c).fill = cor
                        ant = atual
                    else:
                        for c in cols_filha: ws.cell(row=r, column=c).fill = cor

            self.log_gui(f"Arquivo gerado: {self.path_saida.name}")
            self.window.write_event_value("-FINALIZADO-", "SUCESSO")

        except Exception as e:
            self.log_gui(f"ERRO: {str(e)}", "ERROR")
            self.window.write_event_value("-FINALIZADO-", f"ERRO: {str(e)}")

# =============================================================================
# INTERFACE GRÁFICA (UI)
# =============================================================================

def criar_janela():
    sg.theme("SystemDefaultForReal")
    
    layout = [
        [sg.Text("Processamento de Rateio de Faturas", font=("Consolas", 16, "bold"), text_color="#2C3E50")],
        [sg.HorizontalSeparator()],
        [sg.Text("Selecione os arquivos necessários abaixo:", font=("Arial", 10, "italic"))],
        
        [sg.Text("Extrato Sintético:", size=(18, 1)), sg.Input(key="-EXTRATO-"), sg.FileBrowse()],
        [sg.Text("Planilha Inicial:", size=(18, 1)), sg.Input(key="-INICIAL-"), sg.FileBrowse()],
        [sg.Text("Regras de Rateio:", size=(18, 1)), sg.Input(key="-REGRAS-"), sg.FileBrowse()],
        
        [sg.Text("", size=(1, 1))], # Substituição do sg.Space com erro
        
        [sg.Button("Executar", size=(15, 1), button_color=("white", "#27AE60")), 
         sg.Button("Sair", size=(10, 1), button_color=("white", "#C0392B"))],
        
        [sg.Text("Console de Log:", font=("Arial", 10, "bold"))],
        [sg.Multiline(key="-LOG-", size=(80, 15), background_color="black", text_color="#00FF00", 
                      font=("Consolas", 10), autoscroll=True, disabled=True)]
    ]
    
    return sg.Window("Automação de Rateio v1.0", layout, finalize=True)

# =============================================================================
# LOOP PRINCIPAL
# =============================================================================

def main():
    janela = criar_janela()
    log_file = configurar_logger()
    logging.info("Aplicação iniciada.")

    while True:
        evento, valores = janela.read()

        if evento in (sg.WIN_CLOSED, "Sair"):
            break

        if evento == "Executar":
            # Validação
            if not all([valores['-EXTRATO-'], valores['-INICIAL-'], valores['-REGRAS-']]):
                sg.popup_error("Erro", "Por favor, selecione todos os arquivos.")
                continue

            janela["Executar"].update(disabled=True)
            janela["-LOG-"].update("--- INICIANDO PROCESSAMENTO ---\n", append=True)
            
            # Disparo da Thread
            processador = ProcessadorRateio(janela, valores)
            threading.Thread(target=processador.executar, daemon=True).start()

        if evento == "-THREAD_LOG-":
            msg = valores["-THREAD_LOG-"]
            janela["-LOG-"].update(f"{msg}\n", append=True)

        if evento == "-FINALIZADO-":
            janela["Executar"].update(disabled=False)
            resultado = valores["-FINALIZADO-"]
            if resultado == "SUCESSO":
                sg.popup("Sucesso!", "O processamento foi concluído com sucesso.")
            else:
                sg.popup_error("Falha", resultado)

    janela.close()
    logging.info("Aplicação encerrada.")

if __name__ == "__main__":
    main()